# -*- coding: utf-8 -*-
"""Tests for importing the real matching workbook and exporting an identical layout."""
import pytest
from openpyxl import load_workbook

from app.excel_io import import_workbook, export_workbook, HEADERS

REAL_FILE = r"C:\Software\PBK\مطابقة المشتريات والمبيعات.xlsx"

# cached aggregate values from the real workbook (row 4 SUBTOTALs + J2/R2)
SHEET_E4_QTY_PURCH = 515846.66
SHEET_H4_VAT_PURCH = 1148663.943
SHEET_I4_TOTAL_PURCH = 9353406.393
SHEET_N4_QTY_SALES = 530765
SHEET_P4_NET_SALES = 9474187.5
SHEET_Q4_VAT_SALES = 668125.43
SHEET_R4_TOTAL_SALES = 10142312.93


@pytest.fixture(scope="module")
def imported():
    return import_workbook(REAL_FILE)


class TestImportRealWorkbook:
    def test_line_counts(self, imported):
        assert len(imported["purchases"]) == 54   # 58 rows − 4 sales-only
        assert len(imported["sales"]) == 56       # 58 rows − 2 purchase-only
        assert len(imported["matches"]) == 52     # rows holding both sides

    def test_first_purchase_line(self, imported):
        p = imported["purchases"][0]
        assert p["invoice_no"] == "14426"
        assert p["party"] == "سيفتى باك"
        assert p["item"] == "نموذج 2"
        assert p["qty"] == 24000
        assert p["unit_price"] == 9
        assert p["invoice_date"] == "2026-01-04"
        assert p["vat"] == pytest.approx(24000 * 9 * 0.14)

    def test_first_sale_line(self, imported):
        s = imported["sales"][0]
        assert s["party"] == "الفشاوي"
        assert s["qty"] == 24000
        assert s["unit_price"] == 10.5
        assert s["vat"] == 0          # hardcoded 0 in the sheet (unregistered buyer)
        assert s["internal_ref"] == "1012026"
        assert s["invoice_no"] == "19-2026"

    def test_purchase_totals_match_sheet_subtotals(self, imported):
        ps = imported["purchases"]
        assert sum(p["qty"] for p in ps) == pytest.approx(SHEET_E4_QTY_PURCH)
        assert sum(p["vat"] for p in ps) == pytest.approx(SHEET_H4_VAT_PURCH, rel=1e-6)
        total = sum(p["qty"] * p["unit_price"] + p["vat"] for p in ps)
        assert total == pytest.approx(SHEET_I4_TOTAL_PURCH, rel=1e-6)

    def test_sales_totals_match_sheet_subtotals(self, imported):
        ss = imported["sales"]
        assert sum(s["qty"] for s in ss) == pytest.approx(SHEET_N4_QTY_SALES)
        net = sum(s["qty"] * s["unit_price"] for s in ss)
        assert net == pytest.approx(SHEET_P4_NET_SALES, rel=1e-6)
        assert sum(s["vat"] for s in ss) == pytest.approx(SHEET_Q4_VAT_SALES, rel=1e-6)
        total = net + sum(s["vat"] for s in ss)
        assert total == pytest.approx(SHEET_R4_TOTAL_SALES, rel=1e-6)

    def test_text_dates_are_parsed_with_inferred_year(self, imported):
        # row 47: A47='21/4', J47='26/4' — must become 2026 dates
        dated = [p for p in imported["purchases"] if p["invoice_no"] == "849"]
        assert dated and dated[0]["invoice_date"] == "2026-04-21"

    def test_credit_note_annotation_kept(self, imported):
        noted = [s for s in imported["sales"] if s.get("note")]
        assert any("اشعار دائن" in s["note"] for s in noted)

    def test_sale_without_price_warns_not_crashes(self, imported):
        assert isinstance(imported["warnings"], list)


def _line(id, date, no, party, item, qty, price, vat, **kw):
    d = {"id": id, "invoice_date": date, "invoice_no": no, "party": party,
         "item": item, "qty": qty, "unit_price": price, "vat": vat,
         "doc_type": kw.pop("doc_type", "i")}
    d.update(kw)
    return d


class TestExport:
    @pytest.fixture
    def small_export(self, tmp_path):
        purchases = [
            _line(1, "2026-01-04", "14426", "سيفتى باك", "نموذج 2", 24000, 9, 30240.0),
            _line(2, "2026-02-01", "999", "كارتون باك", "نموذج 4", 100, 30, 420.0),
        ]
        sales = [
            _line(11, "2026-01-18", "19-2026", "الفشاوي", "نموذج 2", 24000, 10.5, 0,
                  internal_ref="1012026", note=""),
            _line(12, "2026-02-10", "55", "براميدز", "نموذج 4", 50, 36, 252.0,
                  internal_ref="", note="اشعار دائن"),
        ]
        path = str(tmp_path / "out.xlsx")
        export_workbook(purchases, sales, [(1, 11)], path, vat_rate=0.14)
        return path

    def test_layout_matches_template(self, small_export):
        wb = load_workbook(small_export)
        ws = wb.active
        assert ws.sheet_view.rightToLeft is True
        headers = [ws.cell(row=5, column=c).value for c in range(1, 21)]
        assert headers == HEADERS
        assert ws["J2"].value == "=Q4-H4"
        assert ws["R2"].value == "=R4-I4"
        assert ws["E4"].value == "=SUBTOTAL(9,E6:E8)"
        assert "Table1" in ws.tables
        assert ws.tables["Table1"].ref == "A5:T8"

    def test_row_formulas_and_values(self, small_export):
        wb = load_workbook(small_export)
        ws = wb.active
        # row 6 = matched pair
        assert ws["E6"].value == 24000
        assert ws["G6"].value == "=E6*F6"
        assert ws["H6"].value == "=E6*F6*0.14"   # VAT equals 14% → formula form
        assert ws["I6"].value == "=G6+H6"
        assert ws["Q6"].value == 0               # VAT 0 ≠ 14% → literal value
        assert ws["T6"].value == "=E6-N6"
        # row 7 = unmatched purchase: sale side empty
        assert ws["E7"].value == 100
        assert ws["N7"].value is None
        # row 8 = unmatched sale with credit-note annotation in S
        assert ws["N8"].value == 50
        assert ws["S8"].value == "اشعار دائن"

    def test_round_trip_preserves_data(self, small_export):
        back = import_workbook(small_export)
        assert len(back["purchases"]) == 2
        assert len(back["sales"]) == 2
        assert len(back["matches"]) == 1
        p = back["purchases"][0]
        assert (p["qty"], p["unit_price"], p["vat"]) == (24000, 9, pytest.approx(30240))
        s = back["sales"][1]
        assert s["note"] == "اشعار دائن"
        assert s["vat"] == pytest.approx(252)

    def test_real_file_round_trip_totals(self, imported, tmp_path):
        path = str(tmp_path / "rt.xlsx")
        # assign ids the way the API layer would
        for i, p in enumerate(imported["purchases"]):
            p["id"] = i + 1
        for i, s in enumerate(imported["sales"]):
            s["id"] = 1000 + i + 1
        pairs = [(pi + 1, 1000 + si + 1) for pi, si in imported["matches"]]
        export_workbook(imported["purchases"], imported["sales"], pairs, path, vat_rate=0.14)
        back = import_workbook(path)
        assert len(back["purchases"]) == 54
        assert len(back["sales"]) == 56
        assert len(back["matches"]) == 52
        assert sum(p["vat"] for p in back["purchases"]) == pytest.approx(SHEET_H4_VAT_PURCH, rel=1e-6)
        assert sum(s["vat"] for s in back["sales"]) == pytest.approx(SHEET_Q4_VAT_SALES, rel=1e-6)
        assert sum(s["qty"] for s in back["sales"]) == pytest.approx(SHEET_N4_QTY_SALES)
