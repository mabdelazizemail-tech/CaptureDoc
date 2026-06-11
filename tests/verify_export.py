# -*- coding: utf-8 -*-
"""Verify the exported workbook's formulas evaluate to the expected totals."""
import re
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

PATH = r"C:\Software\PBK\data\shots\exported.xlsx"
wb = load_workbook(PATH)
ws = wb.active

values = {}   # resolved numeric value per coordinate
formulas = {}
for row in ws.iter_rows():
    for c in row:
        if isinstance(c.value, str) and c.value.startswith("="):
            formulas[c.coordinate] = c.value
        elif isinstance(c.value, (int, float)):
            values[c.coordinate] = float(c.value)

CELL = re.compile(r"^[A-T]\d+$")


def resolve(coord, seen=()):
    if coord in values:
        return values[coord]
    if coord not in formulas:
        return 0.0  # empty cell behaves as 0 in Excel arithmetic
    if coord in seen:
        raise ValueError(f"circular ref at {coord}")
    f = formulas[coord][1:]
    m = re.fullmatch(r"SUBTOTAL\(9,([A-T])(\d+):\1(\d+)\)", f)
    if m:
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(3))
        total = sum(resolve(f"{col}{r}", seen + (coord,)) for r in range(r1, r2 + 1))
    else:
        expr = re.sub(r"[A-T]\d+", lambda mm: str(resolve(mm.group(0), seen + (coord,))), f)
        if not re.fullmatch(r"[\d.eE+\-*/() ]+", expr):
            raise ValueError(f"unexpected formula {coord}: {f}")
        total = eval(expr)  # noqa: S307 — sanitized arithmetic only
    values[coord] = total
    return total


bad = []
for coord in list(formulas):
    try:
        resolve(coord)
    except Exception as e:  # noqa: BLE001
        bad.append(f"{coord}: {e}")

print(f"formulas evaluated: {len(formulas)}, problems: {len(bad)}")
for b in bad[:10]:
    print("  ", b)

expect = {
    "H4": 1148663.943, "I4": 9353406.393, "N4": 530765.0,
    "P4": 9474187.5, "Q4": 668125.43, "R4": 10142312.93,
    "E4": 515846.66, "J2": -480538.513, "R2": 788906.537,
}
ok = True
for coord, want in expect.items():
    got = values.get(coord)
    match = got is not None and abs(got - want) < 0.01
    ok &= match
    print(f"{'PASS' if match else 'FAIL'} {coord}: {got:,.3f} (sheet: {want:,.3f})")

table = ws.tables["Table1"].ref if "Table1" in ws.tables else "MISSING"
print("Table1 ref:", table, "| RTL:", ws.sheet_view.rightToLeft)
sys.exit(0 if ok and not bad else 1)
