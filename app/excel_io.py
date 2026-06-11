# -*- coding: utf-8 -*-
"""Import/export of the مطابقة المشتريات والمبيعات workbook.

The export reproduces the user's original template exactly: RTL sheet, headers on
row 5 (including the trailing spaces in 'الكمية ' and 'القيمة المضافة '), Table1
over the data range, SUBTOTAL row 4 and the J2/R2 difference formulas.
"""
from __future__ import annotations

from collections import Counter
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Exact header strings from the original file — trailing spaces are intentional.
HEADERS = [
    "تاريخ فاتورة المشتريات", "رقم الفاتورة", "المورد", "الصنف", "الكمية ",
    "السعر", "Column1", "القيمة المضافة ", "الإجمالي", "تاريخ فاتورة البيع",
    "المشتري", "فاتورة ا/رانيا", "رقم الفاتورة2", "الكمية", "السعر2", "السعر3",
    "القيمة المضافة", "الاجمالي3", "الاجمالي4", "الاجمالي5",
]

NUM_FMT = "#,##0_);[Red](#,##0)"
DATE_FMT = "mm-dd-yy"
DIFF_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'


def _to_float(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _to_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _parse_cell_date(v, fallback_year):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            pass
    # day/month without year, e.g. '21/4' — borrow the sheet's dominant year
    for sep in ("/", "-"):
        parts = s.split(sep)
        if len(parts) == 2:
            try:
                d, m = int(parts[0]), int(parts[1])
                return date(fallback_year, m, d).isoformat()
            except ValueError:
                pass
    return None


def import_workbook(path: str) -> dict:
    wb_f = load_workbook(path)                  # raw values / formula strings
    wb_v = load_workbook(path, data_only=True)  # Excel's cached results
    ws_f, ws_v = wb_f.active, wb_v.active
    for name in wb_f.sheetnames:  # prefer the sheet that holds the matching table
        if any(t.startswith("A5") for t in
               (wb_f[name].tables[k].ref for k in wb_f[name].tables)):
            ws_f, ws_v = wb_f[name], wb_v[name]
            break

    last_row = ws_f.max_row
    for tbl_name in ws_f.tables:
        ref = ws_f.tables[tbl_name].ref  # e.g. 'A5:T63'
        if ref.startswith("A5"):
            last_row = int("".join(ch for ch in ref.split(":")[1] if ch.isdigit()))
            break

    years = Counter()
    for col in ("A", "J"):
        for r in range(6, last_row + 1):
            v = ws_f[f"{col}{r}"].value
            if isinstance(v, datetime):
                years[v.year] += 1
    fallback_year = years.most_common(1)[0][0] if years else date.today().year

    def raw(col, r):
        v = ws_f[f"{col}{r}"].value
        if isinstance(v, str) and v.startswith("="):
            return ws_v[f"{col}{r}"].value  # input cell holding a formula → cached result
        return v if v is not None else ws_v[f"{col}{r}"].value

    purchases, sales, matches, warnings = [], [], [], []
    for r in range(6, last_row + 1):
        item = _to_str(raw("D", r))
        p_qty, p_price = _to_float(raw("E", r)), _to_float(raw("F", r))
        s_qty, s_price = _to_float(raw("N", r)), _to_float(raw("O", r))

        p_idx = s_idx = None
        if p_qty is not None and p_price is not None:
            vat = _to_float(ws_v[f"H{r}"].value)
            if vat is None:
                vat = p_qty * p_price * 0.14
            pdate = _parse_cell_date(raw("A", r), fallback_year)
            if raw("A", r) is not None and pdate is None:
                warnings.append(f"صف {r}: تعذر قراءة تاريخ فاتورة المشتريات «{raw('A', r)}»")
            purchases.append({
                "invoice_date": pdate, "invoice_no": _to_str(raw("B", r)),
                "party": _to_str(raw("C", r)), "item": item,
                "qty": p_qty, "unit_price": p_price, "vat": vat, "doc_type": "i",
            })
            p_idx = len(purchases) - 1

        if s_qty is not None:
            if s_price is None:
                s_price = 0.0
                warnings.append(f"صف {r}: سعر البيع مفقود — تم اعتباره صفراً")
            vat = _to_float(ws_v[f"Q{r}"].value)
            if vat is None:
                q_raw = ws_f[f"Q{r}"].value
                vat = s_qty * s_price * 0.14 if isinstance(q_raw, str) and q_raw.startswith("=") else 0.0
            sdate = _parse_cell_date(raw("J", r), fallback_year)
            if raw("J", r) is not None and sdate is None:
                warnings.append(f"صف {r}: تعذر قراءة تاريخ فاتورة البيع «{raw('J', r)}»")
            sales.append({
                "invoice_date": sdate, "invoice_no": _to_str(raw("M", r)),
                "party": _to_str(raw("K", r)), "item": item,
                "qty": s_qty, "unit_price": s_price, "vat": vat,
                "internal_ref": _to_str(raw("L", r)), "note": _to_str(raw("S", r)),
                "doc_type": "c" if "دائن" in _to_str(raw("S", r)) else "i",
            })
            s_idx = len(sales) - 1

        if p_idx is not None and s_idx is not None:
            matches.append((p_idx, s_idx))

    return {"purchases": purchases, "sales": sales, "matches": matches, "warnings": warnings}


def _smart_number(s):
    s = _to_str(s)
    if s and s.isdigit():
        return int(s)
    return s or None


def _write_date(ws, coord, iso):
    if iso:
        try:
            ws[coord] = datetime.strptime(str(iso)[:10], "%Y-%m-%d")
            ws[coord].number_format = DATE_FMT
        except ValueError:
            ws[coord] = str(iso)


def export_workbook(purchases: list[dict], sales: list[dict],
                    matches: list[tuple], path: str, vat_rate: float = 0.14):
    p_by_id = {p["id"]: p for p in purchases}
    s_by_id = {s["id"]: s for s in sales}
    matched_p = {pid for pid, _ in matches}
    matched_s = {sid for _, sid in matches}

    rows: list[tuple] = [(p_by_id[pid], s_by_id[sid]) for pid, sid in matches
                         if pid in p_by_id and sid in s_by_id]
    rows += [(p, None) for p in purchases if p["id"] not in matched_p]
    rows += [(None, s) for s in sales if s["id"] not in matched_s]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.rightToLeft = True

    rate = f"{vat_rate:g}"
    last = 5 + max(len(rows), 1)

    ws["J2"] = "=Q4-H4"
    ws["R2"] = "=R4-I4"
    for col in ("E", "H", "I", "N", "P", "Q", "R"):
        ws[f"{col}4"] = f"=SUBTOTAL(9,{col}6:{col}{last})"
        ws[f"{col}4"].number_format = NUM_FMT

    header_fill = PatternFill("solid", start_color="D9E1F2")
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=5, column=c, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def vat_cell(coord, qty_col, price_col, r, vat, qty, price):
        expected = (qty or 0) * (price or 0) * vat_rate
        if vat is not None and abs(vat - expected) <= 0.005:
            ws[coord] = f"={qty_col}{r}*{price_col}{r}*{rate}"
        else:
            ws[coord] = vat
        ws[coord].number_format = NUM_FMT

    for i, (p, s) in enumerate(rows):
        r = 6 + i
        item = (p or {}).get("item") or (s or {}).get("item") or ""
        ws[f"D{r}"] = item or None
        if p:
            _write_date(ws, f"A{r}", p.get("invoice_date"))
            ws[f"B{r}"] = _smart_number(p.get("invoice_no"))
            ws[f"C{r}"] = p.get("party") or None
            ws[f"E{r}"] = p.get("qty")
            ws[f"F{r}"] = p.get("unit_price")
            ws[f"G{r}"] = f"=E{r}*F{r}"
            vat_cell(f"H{r}", "E", "F", r, p.get("vat"), p.get("qty"), p.get("unit_price"))
            ws[f"I{r}"] = f"=G{r}+H{r}"
            for col in ("E", "G", "I"):
                ws[f"{col}{r}"].number_format = NUM_FMT
        if s:
            _write_date(ws, f"J{r}", s.get("invoice_date"))
            ws[f"K{r}"] = s.get("party") or None
            ws[f"L{r}"] = _smart_number(s.get("internal_ref"))
            ws[f"M{r}"] = _smart_number(s.get("invoice_no"))
            ws[f"N{r}"] = s.get("qty")
            ws[f"O{r}"] = s.get("unit_price")
            ws[f"P{r}"] = f"=N{r}*O{r}"
            vat_cell(f"Q{r}", "N", "O", r, s.get("vat"), s.get("qty"), s.get("unit_price"))
            ws[f"R{r}"] = f"=P{r}+Q{r}"
            if s.get("note"):
                ws[f"S{r}"] = s["note"]
            ws[f"N{r}"].number_format = NUM_FMT
        ws[f"T{r}"] = f"=E{r}-N{r}"
        ws[f"T{r}"].number_format = DIFF_FMT

    if rows:
        table = Table(displayName="Table1", ref=f"A5:T{5 + len(rows)}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    widths = {"A": 13, "B": 10, "C": 16, "D": 24, "E": 10, "F": 9, "G": 12, "H": 12,
              "I": 13, "J": 13, "K": 20, "L": 12, "M": 11, "N": 10, "O": 9, "P": 12,
              "Q": 12, "R": 13, "S": 12, "T": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for c in range(1, 21):
        ws.cell(row=4, column=c).font = Font(bold=True)

    wb.save(path)
    return path
